import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt

btc_workbook = load_workbook('project1\BTC-USD.xlsx')

btc_sheet = btc_workbook.active
closing_values = np.array([])
for value in btc_sheet.iter_rows(min_row=2,
                                min_col=6,
                                max_col=6,
                                values_only=True):
    closing_values = np.append(closing_values, value)

mean_cv = np.sum(closing_values) / closing_values.size
print(f"Mean of the closing values: {mean_cv}")
alpha = closing_values - mean_cv
variance = np.dot(alpha.T, alpha) / closing_values.size
print(f"Variance of the closing values: {variance}")

num_bins = 10

counts_cv = np.zeros(num_bins)
bins_cv = np.linspace(np.amin(closing_values), np.amax(closing_values), num_bins+1)
for value in closing_values:
    for i in range(num_bins-1, -1, -1):
        if value >= bins_cv[i]:
            counts_cv[i] += 1
            break

plt.stairs(counts_cv, bins_cv, fill=True)
plt.title("Histogram of BTC-USD closing values\n1/1/2017 - 1/7/2023")
plt.xlabel('USD ($)')
plt.ylabel('Count')
plt.show()

plt.stairs((counts_cv / (np.sum(counts_cv) * np.diff(bins_cv))), bins_cv, fill=True)
plt.title("Normalized histogram of BTC-USD closing values\n1/1/2017 - 1/7/2023")
plt.xlabel('USD ($)')
plt.ylabel('Normalized Count')
plt.show()

# using the built in values
plt.hist(closing_values)
plt.title("Histogram of BTC-USD closing values\nusing built in function, 1/1/2017 - 1/7/2023")
plt.xlabel('USD ($)')
plt.ylabel('Count')
plt.show()

plt.hist(closing_values, density=True)
plt.title("Normalized histogram of BTC-USD closing values\nusing built in function, 1/1/2017 - 1/7/2023")
plt.xlabel('USD ($)')
plt.ylabel('Normalized Count')
plt.show()

num_days = np.arange(1, closing_values.size + 1)
plt.plot(num_days, closing_values, 'o', label='Closing Values')
plt.xlabel('Days')
plt.ylabel('USD ($)')
plt.title('BTC-USD closing values, 1/1/2017 - 1/7/2023')
plt.plot(num_days, mean_cv*np.ones((closing_values.size, 1)), 'r', label='Closing Value Mean')
#plt.plot(num_days, np.median(closing_values)*np.ones((closing_values.size, 1)), 'g', label='Closing Value Median')
plt.legend()
plt.show()

x = num_days
X_aug = np.stack((x**5, x**4, x**3, x**2, x, np.ones((x.size))), axis=1)
Beta = np.dot(np.dot(np.linalg.inv(np.dot(X_aug.T, X_aug)), X_aug.T), closing_values)

Y_hat = np.dot(X_aug, Beta)
mean_Y_hat = np.sum(Y_hat) / Y_hat.size
plt.plot(num_days, closing_values, 'o', label='Closing Values')
plt.xlabel('Days')
plt.ylabel('USD ($)')
plt.title('BTC-USD closing values, 1/1/2017 - 1/7/2023')
plt.plot(num_days, mean_cv*np.ones((closing_values.size, 1)), 'r', label='Closing Value Mean')
plt.plot(num_days, Y_hat, 'y', label='LSR Model')
plt.plot(num_days, mean_Y_hat*np.ones((Y_hat.size, 1)), 'g', linestyle='dashed', label='LSR Model Mean')
plt.legend()
plt.show()